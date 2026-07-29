"""
Builds the final unified Excel report containing every flagged student,
their metrics, and their AI-generated warning message.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config


def build_report(flagged_students):
    os.makedirs(os.path.dirname(config.OUTPUT_REPORT_PATH), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Warnings"

    headers = [
        "Student ID", "Student Name", "High Performer?",
        "Semester Avg Grade", "Last Week Avg Grade",
        "Semester Absence %", "Last Week Absence %",
        "Reasons Flagged", "AI Warning Message",
    ]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for student in flagged_students:
        ws.append([
            student.get("student_id", ""),
            student.get("student_name", ""),
            "Yes" if student.get("is_high_performer") else "No",
            student.get("semester_avg_grade", ""),
            student.get("last_week_avg_grade", ""),
            student.get("semester_absence_rate", ""),
            student.get("last_week_absence_rate", ""),
            "; ".join(student.get("reasons", [])),
            student.get("ai_warning", ""),
        ])

    # Column widths for readability
    widths = [12, 20, 14, 16, 16, 16, 16, 45, 60]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(config.OUTPUT_REPORT_PATH)
    return config.OUTPUT_REPORT_PATH
